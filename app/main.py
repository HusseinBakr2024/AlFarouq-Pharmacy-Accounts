from datetime import date, datetime, timedelta
import calendar
from io import BytesIO
from typing import Optional

from fastapi import Depends, FastAPI, HTTPException, Request
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import func, or_, inspect, text
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session, joinedload

from app.database import Base, SessionLocal, engine, get_db
from app.models import (Branch, Employee, Supplier, SalesJournal, SalesLine, PurchaseJournal, PurchaseLine,
    User, Customer, Treasury, Bank, ExpenseItem, OtherAccountItem, OpeningStock, TreasuryDeposit,
    ExpenseJournal, ExpenseLine, ExpenseTreasuryPayment, OtherAccountJournal, OtherAccountLine,
    SupplierClaim, SupplierClaimLine, SupplierPaymentJournal, SupplierPaymentAllocation,
    GeneralCheckJournal, IssuedCheck, NotificationRead)

Base.metadata.create_all(bind=engine)

# Keep existing SQLite installations compatible with additive master-data fields.
with engine.begin() as connection:
    if engine.dialect.name == "sqlite":
        current_sql = connection.execute(text("SELECT sql FROM sqlite_master WHERE type='table' AND name='sales_journals'"))
        row = current_sql.fetchone()
        if row and row[0] and "uq_sales_branch_date" in row[0]:
            connection.execute(text("DROP TABLE IF EXISTS sales_journals_new"))
            connection.execute(text(
                "CREATE TABLE sales_journals_new ("
                "id INTEGER NOT NULL PRIMARY KEY, "
                "journal_no VARCHAR(30) NOT NULL, "
                "journal_date DATE NOT NULL, "
                "branch_id INTEGER NOT NULL, "
                "status VARCHAR(20) NOT NULL, "
                "notes VARCHAR(500), "
                "created_at DATETIME NOT NULL, "
                "updated_at DATETIME NOT NULL, "
                "posted_at DATETIME, "
                "FOREIGN KEY(branch_id) REFERENCES branches(id)"
                ")"
            ))
            connection.execute(text(
                "INSERT INTO sales_journals_new (id, journal_no, journal_date, branch_id, status, notes, created_at, updated_at, posted_at) "
                "SELECT id, journal_no, journal_date, branch_id, status, notes, created_at, updated_at, posted_at FROM sales_journals"
            ))
            connection.execute(text("DROP TABLE sales_journals"))
            connection.execute(text("ALTER TABLE sales_journals_new RENAME TO sales_journals"))
            connection.execute(text("CREATE UNIQUE INDEX IF NOT EXISTS ix_sales_journals_journal_no ON sales_journals (journal_no)"))
            connection.execute(text("CREATE INDEX IF NOT EXISTS ix_sales_journals_journal_date ON sales_journals (journal_date)"))
            connection.execute(text("CREATE INDEX IF NOT EXISTS ix_sales_journals_branch_id ON sales_journals (branch_id)"))
            connection.execute(text("CREATE INDEX IF NOT EXISTS ix_sales_journals_status ON sales_journals (status)"))

    columns={column["name"] for column in inspect(connection).get_columns("other_account_items")}
    if "effect_sign" not in columns:
        connection.execute(text("ALTER TABLE other_account_items ADD COLUMN effect_sign INTEGER NOT NULL DEFAULT 1"))
    purchase_journal_columns={column["name"] for column in inspect(connection).get_columns("purchase_journals")}
    if "notice_type" not in purchase_journal_columns:
        connection.execute(text("ALTER TABLE purchase_journals ADD COLUMN notice_type VARCHAR(40) NOT NULL DEFAULT ''"))
        connection.execute(text("UPDATE purchase_journals SET notice_type = COALESCE((SELECT notice_type FROM purchase_lines WHERE purchase_lines.journal_id = purchase_journals.id AND notice_type <> '' LIMIT 1), '') WHERE entry_type = 'notice'"))
    expense_columns={column["name"] for column in inspect(connection).get_columns("expense_items")}
    if "expense_type" not in expense_columns:
        connection.execute(text("ALTER TABLE expense_items ADD COLUMN expense_type VARCHAR(20) NOT NULL DEFAULT 'operating'"))
    account_columns={column["name"] for column in inspect(connection).get_columns("other_account_items")}
    if "opening_debit" not in account_columns:
        connection.execute(text("ALTER TABLE other_account_items ADD COLUMN opening_debit FLOAT NOT NULL DEFAULT 0"))
    if "opening_credit" not in account_columns:
        connection.execute(text("ALTER TABLE other_account_items ADD COLUMN opening_credit FLOAT NOT NULL DEFAULT 0"))
    journal_columns={column["name"] for column in inspect(connection).get_columns("other_account_journals")}
    if "treasury_id" not in journal_columns:
        connection.execute(text("ALTER TABLE other_account_journals ADD COLUMN treasury_id INTEGER"))

app = FastAPI(title="صيدليات الفاروق")
app.mount("/static", StaticFiles(directory="app/static"), name="static")
templates = Jinja2Templates(directory="app/templates")


def seed_master_data() -> None:
    db = SessionLocal()
    try:
        branches = db.query(Branch).order_by(Branch.id).all()
        if not branches:
            branches = [
                Branch(code="BR-0001", name="فرع الفاروق 1"),
                Branch(code="BR-0002", name="فرع الفاروق 2"),
            ]
            db.add_all(branches)
            db.flush()

        if db.query(Employee).count() == 0:
            first = branches[0]
            second = branches[1] if len(branches) > 1 else branches[0]
            db.add_all([
                Employee(code="EM-0001", name="مستخدم الوردية الأولى", job_title="كاشير", branch_id=first.id),
                Employee(code="EM-0002", name="مستخدم الوردية الثانية", job_title="كاشير", branch_id=first.id),
                Employee(code="EM-0003", name="مستخدم الوردية الأولى - فرع 2", job_title="كاشير", branch_id=second.id),
                Employee(code="EM-0004", name="مستخدم الوردية الثانية - فرع 2", job_title="كاشير", branch_id=second.id),
            ])
        db.commit()
    finally:
        db.close()


seed_master_data()


def render(request: Request, template_name: str, page_title: str, active_page: str, **context):
    notification=None; notification_db=SessionLocal()
    try:
        due=notification_db.query(IssuedCheck).filter(IssuedCheck.status=="issued",IssuedCheck.due_date>=date.today(),IssuedCheck.due_date<=date.today()+timedelta(days=4)).order_by(IssuedCheck.id).all()
        if due:
            key="due-checks:"+",".join(str(x.id) for x in due)
            if not notification_db.query(NotificationRead).filter(NotificationRead.notification_key==key).first():notification={"key":key,"count":len(due),"total":sum(x.amount for x in due)}
    finally:notification_db.close()
    context.update({"request": request, "page_title": page_title, "active_page": active_page,"check_notification":notification})
    return templates.TemplateResponse(request=request, name=template_name, context=context)


def ajax_or_redirect(request: Request, message: str, redirect_url: str, success: bool = True):
    if request.headers.get("x-requested-with") == "XMLHttpRequest":
        return JSONResponse({"success": success, "message": message})
    return RedirectResponse(redirect_url, status_code=303)


def parse_date(value: Optional[str], default: Optional[date] = None) -> Optional[date]:
    if not value:
        return default
    value = str(value).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(value, fmt).date()
        except ValueError:
            continue
    return default


def next_journal_no(db: Session, journal_date: date) -> str:
    prefix = f"SJ-{journal_date.year}-"
    last_no = db.query(func.max(SalesJournal.journal_no)).filter(SalesJournal.journal_no.like(f"{prefix}%")).scalar()
    sequence = int(last_no.rsplit("-", 1)[-1]) + 1 if last_no else 1
    return f"{prefix}{sequence:05d}"


@app.get("/", response_class=HTMLResponse)
def home(request: Request):
    return render(request, "home/index.html", "الرئيسية", "home")


@app.get("/dashboard", response_class=HTMLResponse)
def dashboard(request: Request):
    return render(request, "dashboard/index.html", "لوحة المعلومات", "dashboard")


def next_code(db: Session, model, prefix: str) -> str:
    last = db.query(model).order_by(model.id.desc()).first()
    return f"{prefix}-{((last.id if last else 0)+1):04d}"

SETTINGS = {
 "users": (User,"USR","المستخدمون والصلاحيات",[("username","اسم المستخدم","text"),("full_name","الاسم بالكامل","text"),("role","الدور","text"),("permissions","الصلاحيات","text"),("is_active","نشط","bool")]),
 "employees": (Employee,"EMP","الموظفون",[("name","اسم الموظف","text"),("branch_id","الفرع","branch"),("job_title","الوظيفة","text"),("is_active","نشط","bool")]),
 "branches": (Branch,"BR","الفروع",[("name","اسم الفرع","text"),("is_active","نشط","bool")]),
 "customers": (Customer,"CUS","العملاء",[("name","اسم العميل","text"),("phone","رقم التواصل","text"),("opening_debit","افتتاحي مدين","number"),("opening_credit","افتتاحي دائن","number"),("is_active","نشط","bool")]),
 "suppliers": (Supplier,"SUP","الموردون",[("name","اسم المورد","text"),("phone","رقم التواصل","text"),("opening_debit","افتتاحي مدين","number"),("opening_credit","افتتاحي دائن","number"),("is_active","نشط","bool")]),
 "treasuries": (Treasury,"TRE","الخزائن",[("name","اسم الخزينة","text"),("branch_id","الفرع","branch"),("opening_balance","الرصيد الافتتاحي","number"),("is_active","نشط","bool")]),
 "banks": (Bank,"BNK","البنوك",[("name","اسم البنك","text"),("account_number","رقم الحساب","text"),("opening_balance","الرصيد الافتتاحي","number"),("is_active","نشط","bool")]),
 "expenses": (ExpenseItem,"EXP","بنود المصروفات",[("name","اسم المصروف","text"),("expense_type","نوع المصروف","expense_type"),("is_active","نشط","bool")]),
 "other_accounts": (OtherAccountItem,"ACC","الحسابات الأخرى",[("name","اسم الحساب","text"),("opening_debit","افتتاحي مدين","number"),("opening_credit","افتتاحي دائن","number"),("is_active","نشط","bool")]),
 "opening_stock": (OpeningStock,"STK","المخزون الافتتاحي",[("item_name","اسم الصنف","text"),("branch_id","الفرع","branch"),("quantity","الكمية","number"),("unit_cost","تكلفة الوحدة","number"),("notes","ملاحظات","text")]),
}

@app.get("/settings", response_class=HTMLResponse)
def settings(request: Request, tab: str = "users", search: str = "", edit_id: Optional[int] = None, db: Session = Depends(get_db)):
    if tab not in SETTINGS: tab="users"
    model,prefix,title,fields=SETTINGS[tab]; query=db.query(model)
    searchable=[model.code]+[getattr(model,n) for n,_,kind in fields if kind=="text"]
    if search and searchable: query=query.filter(or_(*[x.ilike(f"%{search}%") for x in searchable]))
    records=query.order_by(model.id.desc()).all(); editing=db.get(model,edit_id) if edit_id else None
    tabs=[{"key":k,"title":v[2]} for k,v in SETTINGS.items()]
    return render(request,"settings/index.html","الإعدادات","settings",tab=tab,tabs=tabs,title=title,fields=fields,records=records,editing=editing,next_value=next_code(db,model,prefix),branches=db.query(Branch).filter(Branch.is_active.is_(True)).order_by(Branch.name).all(),search=search,message=request.query_params.get("message",""))

@app.post("/settings/{tab}/save")
async def save_setting(tab: str, request: Request, db: Session = Depends(get_db)):
    if tab not in SETTINGS: raise HTTPException(404)
    model,prefix,_,fields=SETTINGS[tab]; form=await request.form(); record=db.get(model,int(form.get("id"))) if form.get("id") else model(code=next_code(db,model,prefix))
    for name,_,kind in fields:
        raw=form.get(name)
        if kind=="bool": value=bool(raw)
        elif kind=="number": value=float(raw or 0)
        elif kind=="branch": value=int(raw or 0)
        else: value=str(raw or "").strip()
        setattr(record,name,value)
    if tab=="expenses" and record.expense_type not in {"operating","general"}:
        return RedirectResponse("/settings?tab=expenses&message=يجب اختيار نوع المصروف",303)
    if tab=="other_accounts":
        record.opening_debit=max(record.opening_debit,0); record.opening_credit=max(record.opening_credit,0)
        if record.opening_debit>0: record.opening_credit=0
    if not getattr(record,"id",None): db.add(record)
    try: db.commit()
    except IntegrityError:
        db.rollback(); return RedirectResponse(f"/settings?tab={tab}&message=تعذر الحفظ، تحقق من عدم تكرار البيانات",303)
    return RedirectResponse(f"/settings?tab={tab}&message=تم حفظ البيانات بنجاح",303)

@app.post("/settings/{tab}/{record_id}/delete")
def delete_setting(tab: str, record_id: int, db: Session = Depends(get_db)):
    if tab not in SETTINGS: raise HTTPException(404)
    record=db.get(SETTINGS[tab][0],record_id)
    if record:
        try: db.delete(record); db.commit()
        except IntegrityError: db.rollback(); return RedirectResponse(f"/settings?tab={tab}&message=لا يمكن حذف سجل مرتبط بحركات",303)
    return RedirectResponse(f"/settings?tab={tab}&message=تم حذف البيانات",303)


@app.get("/closing-stock", response_class=HTMLResponse)
def closing_stock(request: Request):
    return render(request, "closing_stock.html", "مخزون آخر الفترة", "closing_stock")


@app.get("/sales", response_class=HTMLResponse)
def sales_entry(
    request: Request,
    edit_id: Optional[int] = None,
    search: str = "",
    branch_id: Optional[str] = None,
    journal_date: Optional[str] = None,
    db: Session = Depends(get_db),
):
    selected_branch_id = int(branch_id) if branch_id and str(branch_id).isdigit() else None
    branches = db.query(Branch).filter(Branch.is_active.is_(True)).order_by(Branch.code).all()
    employees = db.query(Employee).filter(Employee.is_active.is_(True)).order_by(Employee.branch_id, Employee.name).all()
    journal = None
    if edit_id:
        journal = db.query(SalesJournal).options(joinedload(SalesJournal.lines).joinedload(SalesLine.employee),joinedload(SalesJournal.treasury_deposit)).filter(SalesJournal.id == edit_id).first()
        if not journal:
            raise HTTPException(404, "اليومية غير موجودة")

    query = db.query(SalesJournal).options(joinedload(SalesJournal.branch), joinedload(SalesJournal.lines))
    if search:
        query = query.filter(SalesJournal.journal_no.contains(search))
    if selected_branch_id:
        query = query.filter(SalesJournal.branch_id == selected_branch_id)
    if journal_date:
        query = query.filter(SalesJournal.journal_date == parse_date(journal_date))
    results = query.order_by(SalesJournal.journal_date.desc(), SalesJournal.id.desc()).limit(100).all()

    return render(
        request, "sales/index.html", "يومية المبيعات", "sales",
        branches=branches, employees=employees, treasuries=db.query(Treasury).filter(Treasury.is_active.is_(True)).order_by(Treasury.name).all(), journal=journal, results=results,
        is_readonly=bool(journal and journal.status == "posted"),
        today=date.today().strftime("%d/%m/%Y"), selected_branch=selected_branch_id, search=search,
        selected_date=journal_date or "", message=request.query_params.get("message", ""),
    )


@app.post("/sales/save")
async def save_sales(request: Request, db: Session = Depends(get_db)):
    form = await request.form()
    journal_id = int(form.get("journal_id") or 0)
    journal_date = parse_date(str(form.get("journal_date") or ""), date.today())
    branch_id = int(form.get("branch_id") or 0)
    notes = str(form.get("notes") or "").strip()
    treasury_id = int(form.get("treasury_id") or 0)

    if not branch_id or not treasury_id:
        return ajax_or_redirect(request, "يجب اختيار الفرع والخزينة", "/sales?message=يجب اختيار الفرع والخزينة", success=False)
    treasury=db.query(Treasury).filter(Treasury.id==treasury_id,Treasury.is_active.is_(True)).first()
    if not treasury or treasury.branch_id!=branch_id:
        return ajax_or_redirect(request, "الخزينة المختارة لا تتبع الفرع", "/sales?message=الخزينة المختارة لا تتبع الفرع", success=False)

    employee_ids = form.getlist("employee_id")
    shift_values = form.getlist("shift_value")
    discounts = form.getlist("discount")
    net_cash_values = form.getlist("net_cash")
    differences = form.getlist("cash_difference")

    valid_lines = []
    for index, employee_id in enumerate(employee_ids):
        if not employee_id:
            continue
        shift = float(shift_values[index] or 0)
        discount = float(discounts[index] or 0)
        difference = float(differences[index] or 0)
        net_cash = shift - discount + difference
        if shift == 0 and discount == 0 and net_cash == 0 and difference == 0:
            continue
        valid_lines.append((int(employee_id), shift, discount, net_cash, difference))

    if not valid_lines:
        return ajax_or_redirect(request, "أدخل حركة مبيعات واحدة على الأقل", "/sales?message=أدخل حركة مبيعات واحدة على الأقل", success=False)

    valid_employee_ids = {
        row.id for row in db.query(Employee).filter(
            Employee.id.in_([line[0] for line in valid_lines]),
            Employee.branch_id == branch_id,
            Employee.is_active.is_(True),
        ).all()
    }
    if len(valid_employee_ids) != len({line[0] for line in valid_lines}):
        return ajax_or_redirect(request, "يوجد مستخدم غير تابع للفرع المختار", "/sales?message=يوجد مستخدم غير تابع للفرع المختار", success=False)

    if journal_id:
        journal = db.query(SalesJournal).options(joinedload(SalesJournal.lines)).filter(SalesJournal.id == journal_id).first()
        if not journal or journal.status == "posted":
            return ajax_or_redirect(request, "لا يمكن تعديل هذه اليومية", "/sales?message=لا يمكن تعديل هذه اليومية", success=False)
        journal.journal_date = journal_date
        journal.branch_id = branch_id
        journal.notes = notes
        journal.lines.clear()
    else:
        journal = SalesJournal(journal_no=next_journal_no(db, journal_date), journal_date=journal_date, branch_id=branch_id, notes=notes)
        db.add(journal)

    for employee_id, shift, discount, net_cash, difference in valid_lines:
        journal.lines.append(SalesLine(employee_id=employee_id, shift_value=shift, discount=discount, net_cash=net_cash, cash_difference=difference))

    total_net=sum(line[3] for line in valid_lines)
    if journal.treasury_deposit:
        journal.treasury_deposit.treasury_id=treasury_id; journal.treasury_deposit.amount=total_net
    else:
        journal.treasury_deposit=TreasuryDeposit(treasury_id=treasury_id,amount=total_net)

    db.commit()
    return ajax_or_redirect(request, "تم حفظ اليومية بنجاح", "/sales?message=تم حفظ اليومية بنجاح")


@app.get("/review", response_class=HTMLResponse)
def unified_review(request: Request, type: str="sales", search: str="", db: Session=Depends(get_db)):
    allowed={"sales","purchases","expenses","other_accounts","supplier_payments","general_checks"}
    if type not in allowed:type="sales"
    if type=="purchases":
        all_items=db.query(PurchaseJournal).options(joinedload(PurchaseJournal.lines).joinedload(PurchaseLine.supplier)).order_by(PurchaseJournal.journal_date.desc(),PurchaseJournal.id.desc()).all()
    elif type=="expenses":all_items=db.query(ExpenseJournal).options(joinedload(ExpenseJournal.branch),joinedload(ExpenseJournal.lines).joinedload(ExpenseLine.expense_item)).order_by(ExpenseJournal.journal_date.desc(),ExpenseJournal.id.desc()).all()
    elif type=="other_accounts":all_items=db.query(OtherAccountJournal).options(joinedload(OtherAccountJournal.lines).joinedload(OtherAccountLine.account)).order_by(OtherAccountJournal.journal_date.desc(),OtherAccountJournal.id.desc()).all()
    elif type=="supplier_payments":all_items=db.query(SupplierPaymentJournal).options(joinedload(SupplierPaymentJournal.supplier),joinedload(SupplierPaymentJournal.allocations)).order_by(SupplierPaymentJournal.journal_date.desc(),SupplierPaymentJournal.id.desc()).all()
    elif type=="general_checks":all_items=db.query(GeneralCheckJournal).options(joinedload(GeneralCheckJournal.account),joinedload(GeneralCheckJournal.checks)).order_by(GeneralCheckJournal.journal_date.desc(),GeneralCheckJournal.id.desc()).all()
    elif type=="sales":all_items=db.query(SalesJournal).options(joinedload(SalesJournal.branch),joinedload(SalesJournal.lines).joinedload(SalesLine.employee)).order_by(SalesJournal.journal_date.desc(),SalesJournal.id.desc()).all()
    else:all_items=[]
    items=[x for x in all_items if x.status=="draft" and (not search or search.lower() in x.journal_no.lower())]
    totals={"draft":len(items),"lines":sum(len(x.lines) for x in items),"value":sum((x.total_net_cash if type=="sales" else x.total_effect if type in {"purchases","general_checks"} else x.total_amount) for x in items)}
    return render(request,"sales/review.html","مراجعة وترحيل","review",journals=items,totals=totals,selected_type=type,search=search,message=request.query_params.get("message",""))

@app.get("/sales/review")
def sales_review_redirect():
    return RedirectResponse("/review", status_code=303)


@app.post("/sales/{journal_id}/post")
def post_sales(journal_id: int, db: Session = Depends(get_db)):
    journal = db.query(SalesJournal).filter(SalesJournal.id == journal_id).first()
    if not journal:
        raise HTTPException(404, "اليومية غير موجودة")
    if journal.status != "posted":
        journal.status = "posted"
        journal.posted_at = datetime.utcnow()
        db.commit()
    return RedirectResponse("/review?message=تم ترحيل اليومية بنجاح", status_code=303)


@app.get("/reports",response_class=HTMLResponse)
def unified_reports(request:Request,tab:str="sales",branch_id:Optional[int]=None,employee_id:Optional[int]=None,supplier_id:Optional[int]=None,expense_item_id:Optional[int]=None,account_id:Optional[int]=None,expense_type:str="",entry_type:str="all",date_from:Optional[str]=None,date_to:Optional[str]=None,status:str="all",journal_no:str="",document_no:str="",db:Session=Depends(get_db)):
    allowed={"sales","purchases","expenses","other_accounts","supplier_payments"}; tab=tab if tab in allowed else "sales"
    filters={"branch_id":branch_id,"employee_id":employee_id,"supplier_id":supplier_id,"expense_item_id":expense_item_id,"account_id":account_id,"expense_type":expense_type,"entry_type":entry_type,"date_from":date_from or "","date_to":date_to or "","status":status,"journal_no":journal_no,"document_no":document_no}
    account_summary=None
    if tab=="sales":
        q=db.query(SalesLine).join(SalesJournal).options(joinedload(SalesLine.employee),joinedload(SalesLine.journal).joinedload(SalesJournal.branch))
        if branch_id:q=q.filter(SalesJournal.branch_id==branch_id)
        if employee_id:q=q.filter(SalesLine.employee_id==employee_id)
        if journal_no:q=q.filter(SalesJournal.journal_no.contains(journal_no))
        if date_from:q=q.filter(SalesJournal.journal_date>=parse_date(date_from))
        if date_to:q=q.filter(SalesJournal.journal_date<=parse_date(date_to))
        if status in {"draft","posted"}:q=q.filter(SalesJournal.status==status)
        lines=q.order_by(SalesJournal.journal_date.desc(),SalesLine.id.desc()).all(); totals=[("الورديات",len(lines)),("المبيعات",sum(x.shift_value for x in lines)),("الخصومات",sum(x.discount for x in lines)),("فروق الخزينة",sum(x.cash_difference for x in lines)),("صافي النقدية",sum(x.net_cash for x in lines))]
    elif tab=="purchases":
        q=db.query(PurchaseLine).join(PurchaseJournal).options(joinedload(PurchaseLine.supplier),joinedload(PurchaseLine.journal))
        if supplier_id:q=q.filter(PurchaseLine.supplier_id==supplier_id)
        if entry_type in {"purchase","notice"}:q=q.filter(PurchaseJournal.entry_type==entry_type)
        if journal_no:q=q.filter(PurchaseJournal.journal_no.contains(journal_no))
        if document_no:q=q.filter(PurchaseLine.document_no.contains(document_no))
        if date_from:q=q.filter(PurchaseJournal.journal_date>=parse_date(date_from))
        if date_to:q=q.filter(PurchaseJournal.journal_date<=parse_date(date_to))
        if status in {"draft","posted"}:q=q.filter(PurchaseJournal.status==status)
        lines=q.order_by(PurchaseJournal.journal_date.desc(),PurchaseLine.id.desc()).all(); totals=[("الحركات",len(lines)),("القيمة صيدلي",sum(x.pharmacy_value for x in lines)),("القيمة جمهور",sum(x.public_value for x in lines)),("تأثير المورد",sum(x.account_effect for x in lines))]
    elif tab=="expenses":
        q=db.query(ExpenseLine).join(ExpenseJournal).options(joinedload(ExpenseLine.expense_item),joinedload(ExpenseLine.journal).joinedload(ExpenseJournal.branch))
        if branch_id:q=q.filter(ExpenseJournal.branch_id==branch_id)
        if expense_type in {"operating","general"}:q=q.filter(ExpenseJournal.expense_type==expense_type)
        if expense_item_id:q=q.filter(ExpenseLine.expense_item_id==expense_item_id)
        if journal_no:q=q.filter(ExpenseJournal.journal_no.contains(journal_no))
        if date_from:q=q.filter(ExpenseJournal.journal_date>=parse_date(date_from))
        if date_to:q=q.filter(ExpenseJournal.journal_date<=parse_date(date_to))
        if status in {"draft","posted"}:q=q.filter(ExpenseJournal.status==status)
        lines=q.order_by(ExpenseJournal.journal_date.desc(),ExpenseLine.id.desc()).all();totals=[("الحركات",len(lines)),("إجمالي المصروفات",sum(x.amount for x in lines))]
    elif tab=="other_accounts":
        q=db.query(OtherAccountLine).join(OtherAccountJournal).options(joinedload(OtherAccountLine.account),joinedload(OtherAccountLine.journal))
        if account_id:q=q.filter(OtherAccountLine.account_id==account_id)
        if journal_no:q=q.filter(OtherAccountJournal.journal_no.contains(journal_no))
        if date_from:q=q.filter(OtherAccountJournal.journal_date>=parse_date(date_from))
        if date_to:q=q.filter(OtherAccountJournal.journal_date<=parse_date(date_to))
        if status in {"draft","posted"}:q=q.filter(OtherAccountJournal.status==status)
        lines=q.order_by(OtherAccountJournal.journal_date.desc(),OtherAccountLine.id.desc()).all();totals=[("الحركات",len(lines)),("إجمالي المبالغ",sum(x.amount for x in lines))]
        if account_id:
            account=db.get(OtherAccountItem,account_id)
            if account:
                opening=(account.opening_debit or 0)-(account.opening_credit or 0);funding=sum(x.amount for x in lines if x.journal.transaction_type=="funding");withdrawal=sum(x.amount for x in lines if x.journal.transaction_type=="withdrawal")
                account_summary=[("الرصيد الافتتاحي",opening),("إجمالي التمويل",funding),("إجمالي السحب",withdrawal),("الرصيد الحالي",opening+funding-withdrawal)]
    else:lines=[]; totals=[("الحركات",0),("الإجمالي",0),("المرحل",0),("غير المرحل",0)]
    return render(request,"reports/unified.html","التقارير","reports",tab=tab,lines=lines,totals=totals,account_summary=account_summary,filters=filters,branches=db.query(Branch).order_by(Branch.name).all(),employees=db.query(Employee).order_by(Employee.name).all(),suppliers=db.query(Supplier).order_by(Supplier.name).all(),expense_items=db.query(ExpenseItem).order_by(ExpenseItem.name).all(),accounts=db.query(OtherAccountItem).order_by(OtherAccountItem.name).all())

@app.get("/reports/sales")
def sales_report_redirect():
    return RedirectResponse("/reports?tab=sales",status_code=303)

def report_lines_query(db: Session, branch_id=None, employee_id=None, date_from=None, date_to=None, status="all"):
    query = db.query(SalesLine).join(SalesJournal).options(joinedload(SalesLine.employee), joinedload(SalesLine.journal).joinedload(SalesJournal.branch))
    if branch_id:
        query = query.filter(SalesJournal.branch_id == branch_id)
    if employee_id:
        query = query.filter(SalesLine.employee_id == employee_id)
    if date_from:
        query = query.filter(SalesJournal.journal_date >= parse_date(date_from))
    if date_to:
        query = query.filter(SalesJournal.journal_date <= parse_date(date_to))
    if status in {"draft", "posted"}:
        query = query.filter(SalesJournal.status == status)
    return query.order_by(SalesJournal.journal_date, SalesJournal.id, SalesLine.id).all()


@app.get("/reports/sales/export")
def export_sales_report(
    branch_id: Optional[int] = None, employee_id: Optional[int] = None,
    date_from: Optional[str] = None, date_to: Optional[str] = None,
    status: str = "all", db: Session = Depends(get_db),
):
    lines = report_lines_query(db, branch_id, employee_id, date_from, date_to, status)
    wb = Workbook()
    ws = wb.active
    ws.title = "تقرير المبيعات"
    ws.sheet_view.rightToLeft = True
    headers = ["رقم اليومية", "التاريخ", "الفرع", "كود الموظف", "المستخدم", "قيمة الوردية", "الخصم", "صافي النقدية", "فروق الخزينة", "الحالة"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="243B53")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for line in lines:
        ws.append([
            line.journal.journal_no, line.journal.journal_date.strftime("%d/%m/%Y"), line.journal.branch.name,
            line.employee.code, line.employee.name, line.shift_value, line.discount, line.net_cash,
            line.cash_difference, "مرحلة" if line.journal.status == "posted" else "غير مرحلة",
        ])
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
    for column, width in enumerate([18, 14, 22, 15, 28, 16, 14, 16, 16, 14], start=1):
        ws.column_dimensions[get_column_letter(column)].width = width
    ws.freeze_panes = "A2"
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    headers_out = {"Content-Disposition": "attachment; filename=AlFarouq_Sales_Report.xlsx"}
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers_out)



def next_purchase_no(db: Session, d: date) -> str:
    prefix=f"PJ-{d.year}-"; last=db.query(func.max(PurchaseJournal.journal_no)).filter(PurchaseJournal.journal_no.like(f"{prefix}%")).scalar(); seq=int(last.rsplit("-",1)[-1])+1 if last else 1; return f"{prefix}{seq:05d}"

@app.get("/purchases", response_class=HTMLResponse)
def purchases(request: Request, edit_id: Optional[int]=None, search: str="", status: str="all", journal_date:Optional[str]=None, db: Session=Depends(get_db)):
    journal=db.query(PurchaseJournal).options(joinedload(PurchaseJournal.lines).joinedload(PurchaseLine.supplier)).filter(PurchaseJournal.id==edit_id).first() if edit_id else None
    q=db.query(PurchaseJournal).options(joinedload(PurchaseJournal.lines).joinedload(PurchaseLine.supplier))
    if search: q=q.join(PurchaseLine).join(Supplier).filter((PurchaseJournal.journal_no.contains(search)) | (PurchaseLine.document_no.contains(search)) | (Supplier.name.contains(search)))
    if journal_date:q=q.filter(PurchaseJournal.journal_date==parse_date(journal_date))
    if status in {"draft","posted"}: q=q.filter(PurchaseJournal.status==status)
    results=q.order_by(PurchaseJournal.journal_date.desc(),PurchaseJournal.id.desc()).limit(100).all()
    return render(request,"purchases/index.html","يومية المشتريات","purchases",suppliers=db.query(Supplier).filter(Supplier.is_active.is_(True)).order_by(Supplier.name).all(),notice_types=db.query(OtherAccountItem).filter(OtherAccountItem.is_active.is_(True)).order_by(OtherAccountItem.name).all(),journal=journal,results=results,today=date.today().strftime("%d/%m/%Y"),readonly=bool(journal and journal.status=="posted"),search=search,selected_status=status,selected_date=journal_date or "",message=request.query_params.get("message",""))

@app.post("/purchases/save")
async def save_purchases(request: Request, db: Session=Depends(get_db)):
    f=await request.form(); jid=int(f.get("journal_id") or 0); d=parse_date(str(f.get("journal_date") or ""),date.today()); et=str(f.get("entry_type") or "purchase")
    fixed_notice_types={"مرتجع":-1,"لم يصل":-1,"خصم إضافي":-1,"ت. إضافية":1,"غرامة":-1,"أخرى":1}
    journal_notice_type=str(f.get("notice_type") or "").strip() if et=="notice" else ""
    if et not in {"purchase","notice"}: return ajax_or_redirect(request, "اختر نوع اليومية", "/purchases?message=اختر نوع اليومية", success=False)
    if et=="notice" and journal_notice_type not in fixed_notice_types: return ajax_or_redirect(request, "اختر نوع الإشعار", "/purchases?message=اختر نوع الإشعار", success=False)
    supplier_ids=f.getlist("supplier_id"); docs=f.getlist("document_no"); pvs=f.getlist("pharmacy_value"); pubs=f.getlist("public_value"); descs=f.getlist("description"); notes=f.getlist("line_notes")
    rows=[]
    for i,sid in enumerate(supplier_ids):
        if not sid or not docs[i].strip(): continue
        pv=float(pvs[i] or 0); pub=float(pubs[i] or 0); disc=abs(((pv/pub)*100)-100) if pub else 0; nt=""; effect=pv
        if et=="notice":
            nt=journal_notice_type
            if pv <= 0:return ajax_or_redirect(request, "أدخل القيمة صيدلي للإشعار", "/purchases?message=أدخل القيمة صيدلي للإشعار", success=False)
            if pub <= 0:return ajax_or_redirect(request, "أدخل القيمة جمهور للإشعار", "/purchases?message=أدخل القيمة جمهور للإشعار", success=False)
            if not (descs[i].strip() if i<len(descs) else ""):return ajax_or_redirect(request, "أدخل البيان لكل سطر إشعار", "/purchases?message=أدخل البيان لكل سطر إشعار", success=False)
            effect=abs(pv) * fixed_notice_types[nt]
        rows.append((int(sid),docs[i].strip(),nt,pv,pub,max(disc,0),effect,descs[i] if i<len(descs) else "",notes[i] if i<len(notes) else ""))
    if not rows: return ajax_or_redirect(request, "أدخل حركة واحدة على الأقل", "/purchases?message=أدخل حركة واحدة على الأقل", success=False)
    for sid,doc,*_ in rows:
        dup=db.query(PurchaseLine).filter(PurchaseLine.supplier_id==sid,PurchaseLine.document_no==doc,PurchaseLine.entry_type==et)
        if jid: dup=dup.filter(PurchaseLine.journal_id!=jid)
        if dup.first(): return ajax_or_redirect(request, f"رقم المستند {doc} مكرر لنفس المورد", f"/purchases?message=رقم المستند {doc} مكرر لنفس المورد", success=False)
    if jid:
        j=db.query(PurchaseJournal).options(joinedload(PurchaseJournal.lines)).filter(PurchaseJournal.id==jid).first()
        if not j or j.status=="posted": return ajax_or_redirect(request, "لا يمكن تعديل اليومية", "/purchases?message=لا يمكن تعديل اليومية", success=False)
        j.journal_date=d; j.entry_type=et; j.notice_type=journal_notice_type; j.lines.clear()
    else:
        j=PurchaseJournal(journal_no=next_purchase_no(db,d),journal_date=d,entry_type=et,notice_type=journal_notice_type); db.add(j)
    for sid,doc,nt,pv,pub,disc,effect,des,ntes in rows: j.lines.append(PurchaseLine(supplier_id=sid,entry_type=et,document_no=doc,notice_type=nt,pharmacy_value=pv,public_value=pub,discount_percent=disc,account_effect=effect,description=des,notes=ntes))
    db.commit(); return ajax_or_redirect(request, "تم حفظ اليومية بنجاح", "/purchases?message=تم حفظ اليومية بنجاح")

@app.post("/purchases/{journal_id}/post")
def post_purchase(journal_id:int,db:Session=Depends(get_db)):
    j=db.query(PurchaseJournal).filter(PurchaseJournal.id==journal_id).first();
    if not j: raise HTTPException(404,"اليومية غير موجودة")
    j.status="posted"; j.posted_at=datetime.utcnow(); db.commit(); return RedirectResponse("/review?type=purchases&message=تم الترحيل",303)

@app.get("/reports/purchases")
def purchase_report_redirect():
    return RedirectResponse("/reports?tab=purchases",status_code=303)

@app.get("/reports/purchases/export")
def export_purchases(supplier_id:Optional[int]=None,entry_type:str="all",status:str="all",document_no:str="",date_from:Optional[str]=None,date_to:Optional[str]=None,db:Session=Depends(get_db)):
    q=db.query(PurchaseLine).join(PurchaseJournal).options(joinedload(PurchaseLine.supplier),joinedload(PurchaseLine.journal));
    if supplier_id:q=q.filter(PurchaseLine.supplier_id==supplier_id)
    if entry_type in {"purchase","notice"}:q=q.filter(PurchaseLine.entry_type==entry_type)
    if status in {"draft","posted"}:q=q.filter(PurchaseJournal.status==status)
    if document_no:q=q.filter(PurchaseLine.document_no.contains(document_no))
    if date_from:q=q.filter(PurchaseJournal.journal_date>=parse_date(date_from))
    if date_to:q=q.filter(PurchaseJournal.journal_date<=parse_date(date_to))
    lines=q.order_by(PurchaseJournal.journal_date,PurchaseLine.id).all(); wb=Workbook(); ws=wb.active; ws.title="تقرير المشتريات"; ws.sheet_view.rightToLeft=True; headers=["اليومية","التاريخ","النوع","المورد","رقم المستند","نوع الإشعار","صيدلي","جمهور","الخصم %","تأثير الحساب","الوصف","الحالة"]; ws.append(headers)
    for c in ws[1]:c.font=Font(bold=True,color="FFFFFF");c.fill=PatternFill("solid",fgColor="7C3AED");c.alignment=Alignment(horizontal="center")
    for x in lines:ws.append([x.journal.journal_no,x.journal.journal_date.strftime("%d/%m/%Y"),"فاتورة شراء" if x.entry_type=="purchase" else (x.journal.notice_type or x.notice_type or "إشعار"),x.supplier.name,x.document_no,x.notice_type,x.pharmacy_value,x.public_value,x.discount_percent,x.account_effect,x.description,"مرحلة" if x.journal.status=="posted" else "غير مرحلة"] )
    for i,w in enumerate([18,14,12,24,18,18,14,14,12,16,28,14],1):ws.column_dimensions[get_column_letter(i)].width=w
    out=BytesIO();wb.save(out);out.seek(0);return StreamingResponse(out,media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",headers={"Content-Disposition":"attachment; filename=AlFarouq_Purchases_Report.xlsx"})

def next_number(db:Session, model, prefix:str, d:date)->str:
    stem=f"{prefix}-{d.year}-"; last=db.query(func.max(model.journal_no)).filter(model.journal_no.like(f"{stem}%")).scalar(); seq=int(last.rsplit("-",1)[-1])+1 if last else 1; return f"{stem}{seq:05d}"

@app.get("/expenses",response_class=HTMLResponse)
def expenses(request:Request,edit_id:Optional[int]=None,search:str="",status:str="all",journal_date:Optional[str]=None,db:Session=Depends(get_db)):
    journal=db.query(ExpenseJournal).options(joinedload(ExpenseJournal.branch),joinedload(ExpenseJournal.lines).joinedload(ExpenseLine.expense_item),joinedload(ExpenseJournal.treasury_payment)).filter(ExpenseJournal.id==edit_id).first() if edit_id else None
    if edit_id and not journal:raise HTTPException(404,"اليومية غير موجودة")
    q=db.query(ExpenseJournal).options(joinedload(ExpenseJournal.branch),joinedload(ExpenseJournal.lines))
    if search:q=q.filter(ExpenseJournal.journal_no.contains(search))
    if status in {"draft","posted"}:q=q.filter(ExpenseJournal.status==status)
    if journal_date:q=q.filter(ExpenseJournal.journal_date==parse_date(journal_date))
    return render(request,"expenses/index.html","يومية مصروفات الفروع","expenses",journal=journal,results=q.order_by(ExpenseJournal.journal_date.desc(),ExpenseJournal.id.desc()).limit(100).all(),branches=db.query(Branch).filter(Branch.is_active.is_(True)).order_by(Branch.name).all(),treasuries=db.query(Treasury).filter(Treasury.is_active.is_(True)).order_by(Treasury.name).all(),expense_items=db.query(ExpenseItem).filter(ExpenseItem.is_active.is_(True)).order_by(ExpenseItem.name).all(),readonly=bool(journal and journal.status=="posted"),choose_treasury=request.query_params.get("choose_treasury")=="1",today=date.today().strftime("%d/%m/%Y"),search=search,selected_status=status,selected_date=journal_date or "",message=request.query_params.get("message",""))

@app.post("/expenses/save")
async def save_expenses(request:Request,db:Session=Depends(get_db)):
    f=await request.form(); jid=int(f.get("journal_id") or 0); d=parse_date(f.get("journal_date"),date.today()); kind=str(f.get("expense_type") or ""); branch_id=int(f.get("branch_id") or 0) or None
    treasury_id=int(f.get("treasury_id") or 0)
    if kind not in {"operating","general"}:return ajax_or_redirect(request, "اختر نوع المصروف", "/expenses?message=اختر نوع المصروف", success=False)
    if kind=="operating" and not branch_id:return ajax_or_redirect(request, "الفرع مطلوب للمصروف التشغيلي", "/expenses?message=الفرع مطلوب للمصروف التشغيلي", success=False)
    if kind=="general":branch_id=None
    if not treasury_id:return ajax_or_redirect(request, "اختر خزينة الصرف", "/expenses?message=اختر خزينة الصرف", success=False)
    treasury=db.query(Treasury).filter(Treasury.id==treasury_id,Treasury.is_active.is_(True)).first()
    if not treasury:return ajax_or_redirect(request, "اختر خزينة صحيحة", "/expenses?message=اختر خزينة صحيحة", success=False)
    if kind=="operating" and treasury.branch_id!=branch_id:return ajax_or_redirect(request, "الخزينة المختارة لا تتبع الفرع", "/expenses?message=الخزينة المختارة لا تتبع الفرع", success=False)
    ids=f.getlist("expense_item_id"); amounts=f.getlist("amount"); notes=f.getlist("line_notes"); rows=[]
    for i,item_id in enumerate(ids):
        amount=float(amounts[i] or 0)
        if item_id and amount>0:rows.append((int(item_id),amount,notes[i].strip() if i<len(notes) else ""))
    if not rows:return ajax_or_redirect(request, "أدخل حركة مصروف واحدة على الأقل", "/expenses?message=أدخل حركة مصروف واحدة على الأقل", success=False)
    valid={x.id for x in db.query(ExpenseItem).filter(ExpenseItem.id.in_([x[0] for x in rows]),ExpenseItem.expense_type==kind,ExpenseItem.is_active.is_(True)).all()}
    if len(valid)!=len(set(x[0] for x in rows)):return ajax_or_redirect(request, "يوجد حساب لا يطابق نوع المصروف", "/expenses?message=يوجد حساب لا يطابق نوع المصروف", success=False)
    if jid:
        j=db.query(ExpenseJournal).options(joinedload(ExpenseJournal.lines),joinedload(ExpenseJournal.treasury_payment)).filter(ExpenseJournal.id==jid).first()
        if not j or j.status=="posted":return ajax_or_redirect(request, "لا يمكن تعديل اليومية", "/expenses?message=لا يمكن تعديل اليومية", success=False)
        j.journal_date=d;j.expense_type=kind;j.branch_id=branch_id;j.lines.clear()
    else:j=ExpenseJournal(journal_no=next_number(db,ExpenseJournal,"EJ",d),journal_date=d,expense_type=kind,branch_id=branch_id);db.add(j)
    for item_id,amount,note in rows:j.lines.append(ExpenseLine(expense_item_id=item_id,amount=amount,notes=note))
    db.commit();db.refresh(j)
    if j.treasury_payment:j.treasury_payment.treasury_id=treasury.id;j.treasury_payment.amount=j.total_amount
    else:j.treasury_payment=ExpenseTreasuryPayment(treasury_id=treasury.id,amount=j.total_amount)
    db.commit();
    return ajax_or_redirect(request, "تم حفظ اليومية بنجاح", "/expenses?message=تم حفظ اليومية بنجاح")

@app.post("/expenses/{journal_id}/treasury")
async def save_expense_treasury(journal_id:int,request:Request,db:Session=Depends(get_db)):
    f=await request.form();treasury_id=int(f.get("treasury_id") or 0)
    j=db.query(ExpenseJournal).options(joinedload(ExpenseJournal.treasury_payment)).filter(ExpenseJournal.id==journal_id).first();treasury=db.query(Treasury).filter(Treasury.id==treasury_id,Treasury.is_active.is_(True)).first()
    if not j or j.status=="posted":return RedirectResponse("/expenses?message=لا يمكن تعديل حركة خزينة هذه اليومية",303)
    if not treasury:return RedirectResponse(f"/expenses?edit_id={journal_id}&choose_treasury=1&message=اختر خزينة صحيحة",303)
    if j.expense_type=="operating" and treasury.branch_id!=j.branch_id:return RedirectResponse(f"/expenses?edit_id={journal_id}&choose_treasury=1&message=الخزينة المختارة لا تتبع الفرع",303)
    if j.treasury_payment:j.treasury_payment.treasury_id=treasury.id;j.treasury_payment.amount=j.total_amount
    else:j.treasury_payment=ExpenseTreasuryPayment(treasury_id=treasury.id,amount=j.total_amount)
    db.commit();return RedirectResponse(f"/expenses?edit_id={journal_id}&message=تم إنشاء حركة الصرف من الخزينة بنجاح",303)

@app.post("/expenses/{journal_id}/post")
def post_expense(journal_id:int,db:Session=Depends(get_db)):
    j=db.get(ExpenseJournal,journal_id)
    if not j:raise HTTPException(404,"اليومية غير موجودة")
    if j.status!="posted":j.status="posted";j.posted_at=datetime.utcnow();db.commit()
    return RedirectResponse("/review?type=expenses&message=تم الترحيل",303)

@app.get("/other-accounts",response_class=HTMLResponse)
def other_accounts(request:Request,edit_id:Optional[int]=None,search:str="",status:str="all",journal_date:Optional[str]=None,db:Session=Depends(get_db)):
    journal=db.query(OtherAccountJournal).options(joinedload(OtherAccountJournal.lines).joinedload(OtherAccountLine.account)).filter(OtherAccountJournal.id==edit_id).first() if edit_id else None
    if edit_id and not journal:raise HTTPException(404,"اليومية غير موجودة")
    q=db.query(OtherAccountJournal).options(joinedload(OtherAccountJournal.lines))
    if search:q=q.filter(OtherAccountJournal.journal_no.contains(search))
    if status in {"draft","posted"}:q=q.filter(OtherAccountJournal.status==status)
    if journal_date:q=q.filter(OtherAccountJournal.journal_date==parse_date(journal_date))
    return render(request,"other_accounts/index.html","يومية الحسابات الأخرى","other_accounts",journal=journal,results=q.order_by(OtherAccountJournal.journal_date.desc(),OtherAccountJournal.id.desc()).limit(100).all(),accounts=db.query(OtherAccountItem).filter(OtherAccountItem.is_active.is_(True)).order_by(OtherAccountItem.name).all(),treasuries=db.query(Treasury).filter(Treasury.is_active.is_(True)).order_by(Treasury.name).all(),readonly=bool(journal and journal.status=="posted"),today=date.today().strftime("%d/%m/%Y"),search=search,selected_status=status,selected_date=journal_date or "",message=request.query_params.get("message",""))

@app.post("/other-accounts/save")
async def save_other_accounts(request:Request,db:Session=Depends(get_db)):
    f=await request.form();jid=int(f.get("journal_id") or 0);d=parse_date(f.get("journal_date"),date.today());kind=str(f.get("transaction_type") or ""); treasury_id=int(f.get("treasury_id") or 0)
    if kind not in {"funding","withdrawal"}:return ajax_or_redirect(request, "اختر نوع الحركة", "/other-accounts?message=اختر نوع الحركة", success=False)
    if not treasury_id:return ajax_or_redirect(request, "اختر الخزينة", "/other-accounts?message=اختر الخزينة", success=False)
    treasury=db.query(Treasury).filter(Treasury.id==treasury_id,Treasury.is_active.is_(True)).first()
    if not treasury:return ajax_or_redirect(request, "اختر خزينة صحيحة", "/other-accounts?message=اختر خزينة صحيحة", success=False)
    ids=f.getlist("account_id");amounts=f.getlist("amount");descriptions=f.getlist("description");rows=[]
    for i,account_id in enumerate(ids):
        amount=float(amounts[i] or 0)
        if account_id and amount>0:rows.append((int(account_id),amount,descriptions[i].strip() if i<len(descriptions) else ""))
    if not rows:return ajax_or_redirect(request, "أدخل حركة واحدة على الأقل", "/other-accounts?message=أدخل حركة واحدة على الأقل", success=False)
    valid={x.id for x in db.query(OtherAccountItem).filter(OtherAccountItem.id.in_([x[0] for x in rows]),OtherAccountItem.is_active.is_(True)).all()}
    if len(valid)!=len(set(x[0] for x in rows)):return ajax_or_redirect(request, "يوجد حساب غير صالح", "/other-accounts?message=يوجد حساب غير صالح", success=False)
    if jid:
        j=db.query(OtherAccountJournal).options(joinedload(OtherAccountJournal.lines)).filter(OtherAccountJournal.id==jid).first()
        if not j or j.status=="posted":return ajax_or_redirect(request, "لا يمكن تعديل اليومية", "/other-accounts?message=لا يمكن تعديل اليومية", success=False)
        j.journal_date=d;j.transaction_type=kind;j.treasury_id=treasury_id;j.lines.clear()
    else:j=OtherAccountJournal(journal_no=next_number(db,OtherAccountJournal,"OJ",d),journal_date=d,transaction_type=kind,treasury_id=treasury_id);db.add(j)
    for account_id,amount,description in rows:j.lines.append(OtherAccountLine(account_id=account_id,amount=amount,description=description))
    db.commit();return ajax_or_redirect(request, "تم حفظ اليومية بنجاح", "/other-accounts?message=تم حفظ اليومية بنجاح")

@app.post("/other-accounts/{journal_id}/post")
def post_other_account(journal_id:int,db:Session=Depends(get_db)):
    j=db.get(OtherAccountJournal,journal_id)
    if not j:raise HTTPException(404,"اليومية غير موجودة")
    if j.status!="posted":j.status="posted";j.posted_at=datetime.utcnow();db.commit()
    return RedirectResponse("/review?type=other_accounts&message=تم الترحيل",303)

def journal_export(title:str,headers:list,rows:list,filename:str):
    wb=Workbook();ws=wb.active;ws.title=title;ws.sheet_view.rightToLeft=True;ws.append(headers)
    for c in ws[1]:c.font=Font(bold=True,color="FFFFFF");c.fill=PatternFill("solid",fgColor="243B53");c.alignment=Alignment(horizontal="center")
    for row in rows:ws.append(row)
    for column in range(1,len(headers)+1):ws.column_dimensions[get_column_letter(column)].width=22
    out=BytesIO();wb.save(out);out.seek(0);return StreamingResponse(out,media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",headers={"Content-Disposition":f"attachment; filename={filename}"})

@app.get("/reports/expenses/export")
def export_expenses(branch_id:Optional[int]=None,expense_type:str="",expense_item_id:Optional[int]=None,status:str="all",journal_no:str="",date_from:Optional[str]=None,date_to:Optional[str]=None,db:Session=Depends(get_db)):
    q=db.query(ExpenseLine).join(ExpenseJournal).options(joinedload(ExpenseLine.expense_item),joinedload(ExpenseLine.journal).joinedload(ExpenseJournal.branch))
    if branch_id:q=q.filter(ExpenseJournal.branch_id==branch_id)
    if expense_type in {"operating","general"}:q=q.filter(ExpenseJournal.expense_type==expense_type)
    if expense_item_id:q=q.filter(ExpenseLine.expense_item_id==expense_item_id)
    if status in {"draft","posted"}:q=q.filter(ExpenseJournal.status==status)
    if journal_no:q=q.filter(ExpenseJournal.journal_no.contains(journal_no))
    if date_from:q=q.filter(ExpenseJournal.journal_date>=parse_date(date_from))
    if date_to:q=q.filter(ExpenseJournal.journal_date<=parse_date(date_to))
    rows=[[x.journal.journal_no,x.journal.journal_date.strftime("%d/%m/%Y"),"مصروف تشغيلي" if x.journal.expense_type=="operating" else "مصروف عمومي",x.journal.branch.name if x.journal.branch else "—",x.expense_item.name,x.amount,x.notes,"مرحلة" if x.journal.status=="posted" else "غير مرحلة"] for x in q.order_by(ExpenseJournal.journal_date,ExpenseLine.id).all()]
    return journal_export("تقرير المصروفات",["اليومية","التاريخ","النوع","الفرع","الحساب","القيمة","ملاحظات","الحالة"],rows,"AlFarouq_Expenses_Report.xlsx")

@app.get("/reports/other-accounts/export")
def export_other_accounts(account_id:Optional[int]=None,status:str="all",journal_no:str="",date_from:Optional[str]=None,date_to:Optional[str]=None,db:Session=Depends(get_db)):
    q=db.query(OtherAccountLine).join(OtherAccountJournal).options(joinedload(OtherAccountLine.account),joinedload(OtherAccountLine.journal))
    if account_id:q=q.filter(OtherAccountLine.account_id==account_id)
    if status in {"draft","posted"}:q=q.filter(OtherAccountJournal.status==status)
    if journal_no:q=q.filter(OtherAccountJournal.journal_no.contains(journal_no))
    if date_from:q=q.filter(OtherAccountJournal.journal_date>=parse_date(date_from))
    if date_to:q=q.filter(OtherAccountJournal.journal_date<=parse_date(date_to))
    rows=[[x.journal.journal_no,x.journal.journal_date.strftime("%d/%m/%Y"),"تمويل" if x.journal.transaction_type=="funding" else "سحب أول",x.account.name,x.amount,x.description,"مرحلة" if x.journal.status=="posted" else "غير مرحلة"] for x in q.order_by(OtherAccountJournal.journal_date,OtherAccountLine.id).all()]
    return journal_export("تقرير الحسابات الأخرى",["اليومية","التاريخ","نوع الحركة","الحساب","المبلغ","البيان","الحالة"],rows,"AlFarouq_Other_Accounts_Report.xlsx")

def treasury_available_balance(db:Session,treasury_id:int)->float:
    treasury=db.get(Treasury,treasury_id)
    if not treasury:return 0
    sales=db.query(func.coalesce(func.sum(TreasuryDeposit.amount),0)).join(SalesJournal).filter(TreasuryDeposit.treasury_id==treasury_id,SalesJournal.status=="posted").scalar() or 0
    expenses=db.query(func.coalesce(func.sum(ExpenseTreasuryPayment.amount),0)).join(ExpenseJournal).filter(ExpenseTreasuryPayment.treasury_id==treasury_id,ExpenseJournal.status=="posted").scalar() or 0
    payments=db.query(func.coalesce(func.sum(SupplierPaymentJournal.total_amount),0)).filter(SupplierPaymentJournal.treasury_id==treasury_id,SupplierPaymentJournal.payment_method=="cash",SupplierPaymentJournal.status=="posted").scalar() or 0
    return (treasury.opening_balance or 0)+sales-expenses-payments

def purchase_movement_name(line:PurchaseLine)->str:
    if line.entry_type=="purchase":return "فاتورة شراء"
    notice=line.notice_type or (line.journal.notice_type if line.journal else "")
    if notice=="مرتجع":return "إشعار مرتجع"
    if notice in {"خصم إضافي","لم يصل","غرامة"}:return "إشعار خصم"
    if notice=="ت. إضافية":return "تكلفة إضافية"
    return "إشعار"

def next_claim_no(db:Session,d:date)->str:
    prefix=f"CLM-{d.year}-";last=db.query(func.max(SupplierClaim.claim_no)).filter(SupplierClaim.claim_no.like(f"{prefix}%")).scalar();seq=int(last.rsplit("-",1)[-1])+1 if last else 1;return f"{prefix}{seq:05d}"

@app.get("/supplier-claims",response_class=HTMLResponse)
def supplier_claims(request:Request,supplier_id:Optional[int]=None,date_from:Optional[str]=None,date_to:Optional[str]=None,db:Session=Depends(get_db)):
    movements=[];balance=0;invoice_count=0;pharmacy_total=0;public_total=0;avg_discount=0;selected_supplier_label=""
    date_from_date=parse_date(date_from) if date_from else None
    date_to_date=parse_date(date_to) if date_to else None
    if supplier_id:
        supplier=db.get(Supplier,supplier_id)
        if supplier:
            selected_supplier_label = f"{supplier.code} — {supplier.name}"
            balance=(supplier.opening_credit or 0)-(supplier.opening_debit or 0)
            claimed=db.query(SupplierClaimLine.purchase_line_id)
            q=db.query(PurchaseLine).join(PurchaseJournal).options(joinedload(PurchaseLine.journal)).filter(
                PurchaseLine.supplier_id==supplier_id,
                PurchaseJournal.status=="posted",
                ~PurchaseLine.id.in_(claimed),
            )
            if date_from_date:
                q=q.filter(PurchaseJournal.journal_date>=date_from_date)
            if date_to_date:
                q=q.filter(PurchaseJournal.journal_date<=date_to_date)
            lines=q.order_by(PurchaseJournal.journal_date,PurchaseLine.document_no,PurchaseLine.id).all()
            for line in lines:
                value=line.account_effect or 0
                balance += value
                discount = line.discount_percent if line.discount_percent is not None else 0
                if discount == 0 and line.public_value:
                    discount = abs(((line.pharmacy_value or 0) / (line.public_value or 1) * 100) - 100)
                movements.append({
                    "line": line,
                    "type": purchase_movement_name(line),
                    "pharmacy_value": line.pharmacy_value or 0,
                    "public_value": line.public_value or 0,
                    "avg_discount": discount,
                    "notes": line.journal.notes or "",
                    "value": value,
                })
            invoice_count = len(movements)
            pharmacy_total = sum(x["pharmacy_value"] for x in movements)
            public_total = sum(x["public_value"] for x in movements)
            avg_discount = round(sum(x["avg_discount"] for x in movements) / invoice_count, 2) if invoice_count else 0
    claims=db.query(SupplierClaim).options(joinedload(SupplierClaim.supplier)).order_by(SupplierClaim.claim_date.desc(),SupplierClaim.id.desc()).limit(100).all()
    return render(request,"supplier/claims.html","مراجعة مطالبات المورد","supplier_claims",
        suppliers=db.query(Supplier).filter(Supplier.is_active.is_(True)).order_by(Supplier.name).all(),
        supplier_id=supplier_id,
        selected_supplier_label=selected_supplier_label,
        date_from_value=date_from_date.isoformat() if date_from_date else "",
        date_to_value=date_to_date.isoformat() if date_to_date else "",
        invoice_count=invoice_count,
        pharmacy_total=pharmacy_total,
        public_total=public_total,
        avg_discount=avg_discount,
        movements=movements,
        claims=claims,
        today=date.today().strftime("%d/%m/%Y"),
        message=request.query_params.get("message",""),
    )

@app.post("/supplier-claims/create")
async def create_supplier_claim(request:Request,db:Session=Depends(get_db)):
    f=await request.form();supplier_id=int(f.get("supplier_id") or 0);ids={int(x) for x in f.getlist("purchase_line_id") if str(x).isdigit()};d=parse_date(f.get("claim_date"),date.today())
    if not supplier_id or not ids:return RedirectResponse(f"/supplier-claims?supplier_id={supplier_id}&message=اختر حركة واحدة على الأقل",303)
    already={x[0] for x in db.query(SupplierClaimLine.purchase_line_id).filter(SupplierClaimLine.purchase_line_id.in_(ids)).all()}
    lines=db.query(PurchaseLine).join(PurchaseJournal).filter(PurchaseLine.id.in_(ids),PurchaseLine.supplier_id==supplier_id,PurchaseJournal.status=="posted").all()
    if already or len(lines)!=len(ids):return RedirectResponse(f"/supplier-claims?supplier_id={supplier_id}&message=توجد حركة غير متاحة أو تمت المطالبة بها",303)
    total=sum(x.account_effect or 0 for x in lines)
    if total<=0:return RedirectResponse(f"/supplier-claims?supplier_id={supplier_id}&message=إجمالي المطالبة يجب أن يكون أكبر من صفر",303)
    claim=SupplierClaim(claim_no=next_claim_no(db,d),claim_date=d,supplier_id=supplier_id,total_amount=total,status="ready");db.add(claim)
    for line in lines:claim.lines.append(SupplierClaimLine(purchase_line_id=line.id,amount=line.account_effect or 0))
    try:db.commit()
    except IntegrityError:db.rollback();return RedirectResponse(f"/supplier-claims?supplier_id={supplier_id}&message=تعذر الاعتماد لأن إحدى الحركات مرتبطة بمطالبة أخرى",303)
    return RedirectResponse(f"/supplier-claims?supplier_id={supplier_id}&message=تم اعتماد المطالبة وأصبحت جاهزة للسداد",303)

@app.get("/supplier-payments",response_class=HTMLResponse)
def supplier_payments(request:Request,supplier_id:Optional[int]=None,db:Session=Depends(get_db)):
    claims=db.query(SupplierClaim).filter(SupplierClaim.supplier_id==supplier_id,SupplierClaim.status=="ready").order_by(SupplierClaim.claim_date,SupplierClaim.claim_no).all() if supplier_id else []
    treasuries=db.query(Treasury).filter(Treasury.is_active.is_(True)).order_by(Treasury.name).all()
    balances={x.id:treasury_available_balance(db,x.id) for x in treasuries}
    journals=db.query(SupplierPaymentJournal).options(joinedload(SupplierPaymentJournal.supplier)).order_by(SupplierPaymentJournal.journal_date.desc(),SupplierPaymentJournal.id.desc()).limit(100).all()
    return render(request,"supplier/payments.html","سداد الموردين","supplier_payments",suppliers=db.query(Supplier).filter(Supplier.is_active.is_(True)).order_by(Supplier.name).all(),supplier_id=supplier_id,claims=claims,treasuries=treasuries,treasury_balances=balances,banks=db.query(Bank).filter(Bank.is_active.is_(True)).order_by(Bank.name).all(),journals=journals,today=date.today().strftime("%d/%m/%Y"),message=request.query_params.get("message",""))

@app.post("/supplier-payments/save")
async def save_supplier_payment(request:Request,db:Session=Depends(get_db)):
    f=await request.form();supplier_id=int(f.get("supplier_id") or 0);method=str(f.get("payment_method") or "");claim_ids={int(x) for x in f.getlist("claim_id") if str(x).isdigit()};d=parse_date(f.get("journal_date"),date.today())
    claims=db.query(SupplierClaim).filter(SupplierClaim.id.in_(claim_ids),SupplierClaim.supplier_id==supplier_id,SupplierClaim.status=="ready").all() if claim_ids else []
    if not supplier_id or method not in {"cash","checks"} or len(claims)!=len(claim_ids):return RedirectResponse(f"/supplier-payments?supplier_id={supplier_id}&message=اختر المطالبات وطريقة السداد",303)
    total=sum(x.remaining for x in claims);treasury_id=int(f.get("treasury_id") or 0) or None
    if method=="cash":
        if not treasury_id or treasury_available_balance(db,treasury_id)<total:return RedirectResponse(f"/supplier-payments?supplier_id={supplier_id}&message=رصيد الخزينة غير كافٍ أو لم يتم اختيار خزينة",303)
    journal=SupplierPaymentJournal(journal_no=next_number(db,SupplierPaymentJournal,"SPJ",d),journal_date=d,supplier_id=supplier_id,payment_method=method,treasury_id=treasury_id,total_amount=total);db.add(journal)
    for claim in claims:journal.allocations.append(SupplierPaymentAllocation(claim_id=claim.id,amount=claim.remaining))
    if method=="checks":
        banks=f.getlist("bank_id");numbers=f.getlist("check_no");amounts=f.getlist("check_amount");dues=f.getlist("due_date");beneficiaries=f.getlist("beneficiary");descriptions=f.getlist("check_description");check_total=0;count=0
        for i,bank_id in enumerate(banks[:10]):
            amount=float(amounts[i] or 0) if i<len(amounts) else 0
            if not bank_id and not (numbers[i].strip() if i<len(numbers) else ""):continue
            if not bank_id or not numbers[i].strip() or amount<=0 or not parse_date(dues[i] if i<len(dues) else ""):return RedirectResponse(f"/supplier-payments?supplier_id={supplier_id}&message=أكمل بيانات جميع الشيكات",303)
            journal.checks.append(IssuedCheck(supplier_id=supplier_id,bank_id=int(bank_id),check_no=numbers[i].strip(),amount=amount,due_date=parse_date(dues[i]),beneficiary=(beneficiaries[i].strip() if i<len(beneficiaries) else "") or db.get(Supplier,supplier_id).name,description=descriptions[i].strip() if i<len(descriptions) else ""));check_total+=amount;count+=1
        if not count or abs(check_total-total)>.01:return RedirectResponse(f"/supplier-payments?supplier_id={supplier_id}&message=إجمالي الشيكات يجب أن يساوي إجمالي المطالبات",303)
    try:db.commit()
    except IntegrityError:db.rollback();return RedirectResponse(f"/supplier-payments?supplier_id={supplier_id}&message=رقم شيك مكرر لنفس البنك",303)
    return RedirectResponse("/supplier-payments?message=تم حفظ اليومية وإرسالها للمراجعة والترحيل",303)

@app.post("/supplier-payments/{journal_id}/post")
def post_supplier_payment(journal_id:int,db:Session=Depends(get_db)):
    journal=db.query(SupplierPaymentJournal).options(joinedload(SupplierPaymentJournal.allocations).joinedload(SupplierPaymentAllocation.claim),joinedload(SupplierPaymentJournal.checks)).filter(SupplierPaymentJournal.id==journal_id).first()
    if not journal:raise HTTPException(404,"اليومية غير موجودة")
    if journal.status=="posted":return RedirectResponse("/review?type=supplier_payments",303)
    if journal.payment_method=="cash" and treasury_available_balance(db,journal.treasury_id)<journal.total_amount:return RedirectResponse("/review?type=supplier_payments&message=تعذر الترحيل: رصيد الخزينة غير كافٍ",303)
    for allocation in journal.allocations:
        allocation.claim.paid_amount=min(allocation.claim.total_amount,allocation.claim.paid_amount+allocation.amount);allocation.claim.status="closed";allocation.claim.closed_at=datetime.utcnow()
    for check in journal.checks:check.status="issued";check.posted_at=datetime.utcnow()
    journal.status="posted";journal.posted_at=datetime.utcnow();db.commit();return RedirectResponse("/review?type=supplier_payments&message=تم ترحيل يومية السداد",303)

@app.get("/general-checks",response_class=HTMLResponse)
def general_checks(request:Request,db:Session=Depends(get_db)):
    return render(request,"checks/general.html","تحرير شيكات عامة","general_checks",accounts=db.query(OtherAccountItem).filter(OtherAccountItem.is_active.is_(True)).order_by(OtherAccountItem.name).all(),banks=db.query(Bank).filter(Bank.is_active.is_(True)).order_by(Bank.name).all(),today=date.today().strftime("%d/%m/%Y"),message=request.query_params.get("message",""))

@app.post("/general-checks/save")
async def save_general_checks(request:Request,db:Session=Depends(get_db)):
    f=await request.form();account_id=int(f.get("account_id") or 0);d=parse_date(f.get("journal_date"),date.today());banks=f.getlist("bank_id");numbers=f.getlist("check_no");amounts=f.getlist("check_amount");dues=f.getlist("due_date");beneficiaries=f.getlist("beneficiary");descriptions=f.getlist("check_description")
    if not db.get(OtherAccountItem,account_id):return RedirectResponse("/general-checks?message=اختر الحساب",303)
    journal=GeneralCheckJournal(journal_no=next_number(db,GeneralCheckJournal,"GCJ",d),journal_date=d,account_id=account_id);db.add(journal);count=0
    for i,bank_id in enumerate(banks[:10]):
        amount=float(amounts[i] or 0) if i<len(amounts) else 0
        if not bank_id and not (numbers[i].strip() if i<len(numbers) else ""):continue
        due=parse_date(dues[i] if i<len(dues) else "")
        if not bank_id or not numbers[i].strip() or amount<=0 or not due or not beneficiaries[i].strip():return RedirectResponse("/general-checks?message=أكمل بيانات جميع الشيكات",303)
        journal.checks.append(IssuedCheck(general_account_id=account_id,bank_id=int(bank_id),check_no=numbers[i].strip(),amount=amount,due_date=due,beneficiary=beneficiaries[i].strip(),description=descriptions[i].strip() if i<len(descriptions) else ""));count+=1
    if not count:return RedirectResponse("/general-checks?message=أدخل شيكاً واحداً على الأقل",303)
    try:db.commit()
    except IntegrityError:db.rollback();return RedirectResponse("/general-checks?message=رقم شيك مكرر لنفس البنك",303)
    return RedirectResponse("/general-checks?message=تم حفظ يومية الشيكات وإرسالها للمراجعة",303)

@app.post("/general-checks/{journal_id}/post")
def post_general_checks(journal_id:int,db:Session=Depends(get_db)):
    journal=db.query(GeneralCheckJournal).options(joinedload(GeneralCheckJournal.checks)).filter(GeneralCheckJournal.id==journal_id).first()
    if not journal:raise HTTPException(404,"اليومية غير موجودة")
    if journal.status!="posted":
        journal.status="posted";journal.posted_at=datetime.utcnow()
        for check in journal.checks:check.status="issued";check.posted_at=datetime.utcnow()
        db.commit()
    return RedirectResponse("/review?type=general_checks&message=تم ترحيل الشيكات المحررة",303)

@app.get("/checks-calendar",response_class=HTMLResponse)
def checks_calendar(request:Request,year:Optional[int]=None,month:Optional[int]=None,selected_date:Optional[str]=None,db:Session=Depends(get_db)):
    today=date.today();year=year or today.year;month=month or today.month
    if month<1 or month>12:month=today.month
    first=date(year,month,1);last=date(year,month,calendar.monthrange(year,month)[1]);checks=db.query(IssuedCheck).options(joinedload(IssuedCheck.bank)).filter(IssuedCheck.status.in_(["issued","cleared"]),IssuedCheck.due_date>=first,IssuedCheck.due_date<=last).order_by(IssuedCheck.due_date,IssuedCheck.check_no).all();by_day={}
    for check in checks:by_day.setdefault(check.due_date.day,[]).append(check)
    weeks=calendar.Calendar(firstweekday=5).monthdatescalendar(year,month);selected=parse_date(selected_date);day_checks=[x for x in checks if selected and x.due_date==selected]
    prev=(first-timedelta(days=1));nxt=(last+timedelta(days=1))
    return render(request,"checks/calendar.html","أجندة الشيكات","checks_calendar",year=year,month=month,weeks=weeks,by_day=by_day,selected=selected,day_checks=day_checks,prev=prev,next=nxt,today=today)

@app.get("/notifications/checks/read")
def read_check_notification(key:str,db:Session=Depends(get_db)):
    if key and not db.query(NotificationRead).filter(NotificationRead.notification_key==key).first():db.add(NotificationRead(notification_key=key));db.commit()
    return RedirectResponse("/checks-calendar",303)

SUPPLIER_REPORT_TABS={"claims":"تقرير مطالبات الموردين","cash":"تقرير المدفوعات النقدية","issued":"تقرير الشيكات المحررة","due":"تقرير الشيكات المستحقة","cleared":"تقرير الشيكات المصروفة","supplier_statement":"كشف حساب مورد شامل","bank":"كشف حركة بنك","treasury":"كشف حركة خزينة"}

def supplier_report_data(db:Session,tab:str,supplier_id=None,bank_id=None,treasury_id=None,date_from=None,date_to=None):
    start=parse_date(date_from);end=parse_date(date_to);rows=[]
    if tab=="claims":
        q=db.query(SupplierClaim).options(joinedload(SupplierClaim.supplier));q=q.filter(SupplierClaim.supplier_id==supplier_id) if supplier_id else q
        if start:q=q.filter(SupplierClaim.claim_date>=start)
        if end:q=q.filter(SupplierClaim.claim_date<=end)
        columns=["رقم المطالبة","التاريخ","المورد","الإجمالي","المسدد","المتبقي","الحالة"]
        rows=[[x.claim_no,x.claim_date.strftime("%d/%m/%Y"),x.supplier.name,x.total_amount,x.paid_amount,x.remaining,"مغلقة" if x.status=="closed" else "جاهزة للسداد"] for x in q.order_by(SupplierClaim.claim_date,SupplierClaim.id).all()]
    elif tab=="cash":
        q=db.query(SupplierPaymentJournal).options(joinedload(SupplierPaymentJournal.supplier),joinedload(SupplierPaymentJournal.treasury)).filter(SupplierPaymentJournal.payment_method=="cash");q=q.filter(SupplierPaymentJournal.supplier_id==supplier_id) if supplier_id else q
        if treasury_id:q=q.filter(SupplierPaymentJournal.treasury_id==treasury_id)
        if start:q=q.filter(SupplierPaymentJournal.journal_date>=start)
        if end:q=q.filter(SupplierPaymentJournal.journal_date<=end)
        columns=["اليومية","التاريخ","المورد","الخزينة","القيمة","الحالة"];rows=[[x.journal_no,x.journal_date.strftime("%d/%m/%Y"),x.supplier.name,x.treasury.name if x.treasury else "—",x.total_amount,"مرحل" if x.status=="posted" else "غير مرحل"] for x in q.order_by(SupplierPaymentJournal.journal_date,SupplierPaymentJournal.id).all()]
    elif tab in {"issued","due","cleared"}:
        q=db.query(IssuedCheck).options(joinedload(IssuedCheck.bank),joinedload(IssuedCheck.supplier),joinedload(IssuedCheck.general_account));q=q.filter(IssuedCheck.supplier_id==supplier_id) if supplier_id else q;q=q.filter(IssuedCheck.bank_id==bank_id) if bank_id else q
        if tab=="issued":q=q.filter(IssuedCheck.status.in_(["issued","cleared"]))
        elif tab=="due":q=q.filter(IssuedCheck.status=="issued",IssuedCheck.due_date>=date.today())
        else:q=q.filter(IssuedCheck.status=="cleared")
        if start:q=q.filter(IssuedCheck.due_date>=start)
        if end:q=q.filter(IssuedCheck.due_date<=end)
        columns=["رقم الشيك","تاريخ الاستحقاق","البنك","المستفيد","الجهة","القيمة","الحالة"];rows=[[x.check_no,x.due_date.strftime("%d/%m/%Y"),x.bank.name,x.beneficiary,x.supplier.name if x.supplier else x.general_account.name if x.general_account else "—",x.amount,"مصروف" if x.status=="cleared" else "محرر"] for x in q.order_by(IssuedCheck.due_date,IssuedCheck.check_no).all()]
    elif tab=="supplier_statement":
        columns=["التاريخ","المرجع","نوع الحركة","مدين","دائن","الرصيد"];balance=0;events=[]
        if supplier_id:
            supplier=db.get(Supplier,supplier_id);balance=(supplier.opening_credit or 0)-(supplier.opening_debit or 0) if supplier else 0
            q=db.query(PurchaseLine).join(PurchaseJournal).options(joinedload(PurchaseLine.journal)).filter(PurchaseLine.supplier_id==supplier_id,PurchaseJournal.status=="posted")
            for x in q.all():events.append((x.journal.journal_date,x.document_no,purchase_movement_name(x),max(-(x.account_effect or 0),0),max(x.account_effect or 0,0)))
            p=db.query(SupplierPaymentJournal).filter(SupplierPaymentJournal.supplier_id==supplier_id,SupplierPaymentJournal.status=="posted")
            for x in p.all():events.append((x.journal_date,x.journal_no,"سداد مورد",x.total_amount,0))
            for d,ref,kind,debit,credit in sorted(events,key=lambda x:(x[0],x[1])):
                if start and d<start:continue
                if end and d>end:continue
                balance+=credit-debit;rows.append([d.strftime("%d/%m/%Y"),ref,kind,debit,credit,balance])
    elif tab=="bank":
        columns=["التاريخ","البنك","رقم الشيك","المستفيد","القيمة","الحالة"];q=db.query(IssuedCheck).options(joinedload(IssuedCheck.bank)).filter(IssuedCheck.status.in_(["issued","cleared"]));q=q.filter(IssuedCheck.bank_id==bank_id) if bank_id else q
        if start:q=q.filter(IssuedCheck.due_date>=start)
        if end:q=q.filter(IssuedCheck.due_date<=end)
        rows=[[x.due_date.strftime("%d/%m/%Y"),x.bank.name,x.check_no,x.beneficiary,x.amount,"مصروف" if x.status=="cleared" else "محرر - لم يخصم"] for x in q.order_by(IssuedCheck.due_date).all()]
    else:
        columns=["التاريخ","الخزينة","المرجع","نوع الحركة","وارد","صادر"]
        treasuries=db.query(Treasury).filter(Treasury.id==treasury_id).all() if treasury_id else db.query(Treasury).all()
        for treasury in treasuries:
            for x in db.query(TreasuryDeposit).join(SalesJournal).filter(TreasuryDeposit.treasury_id==treasury.id,SalesJournal.status=="posted").all():rows.append([x.sales_journal.journal_date.strftime("%d/%m/%Y"),treasury.name,x.sales_journal.journal_no,"مبيعات",x.amount,0])
            for x in db.query(ExpenseTreasuryPayment).join(ExpenseJournal).filter(ExpenseTreasuryPayment.treasury_id==treasury.id,ExpenseJournal.status=="posted").all():rows.append([x.journal.journal_date.strftime("%d/%m/%Y"),treasury.name,x.journal.journal_no,"مصروف",0,x.amount])
            for x in db.query(SupplierPaymentJournal).filter(SupplierPaymentJournal.treasury_id==treasury.id,SupplierPaymentJournal.status=="posted",SupplierPaymentJournal.payment_method=="cash").all():rows.append([x.journal_date.strftime("%d/%m/%Y"),treasury.name,x.journal_no,"سداد مورد",0,x.total_amount])
        rows.sort(key=lambda x:datetime.strptime(x[0],"%d/%m/%Y"))
    return columns,rows

@app.get("/supplier-reports",response_class=HTMLResponse)
def supplier_reports(request:Request,tab:str="claims",supplier_id:Optional[int]=None,bank_id:Optional[int]=None,treasury_id:Optional[int]=None,date_from:Optional[str]=None,date_to:Optional[str]=None,db:Session=Depends(get_db)):
    tab=tab if tab in SUPPLIER_REPORT_TABS else "claims";columns,rows=supplier_report_data(db,tab,supplier_id,bank_id,treasury_id,date_from,date_to)
    return render(request,"reports/supplier_cycle.html",SUPPLIER_REPORT_TABS[tab],"supplier_reports",tab=tab,tabs=SUPPLIER_REPORT_TABS,columns=columns,rows=rows,suppliers=db.query(Supplier).order_by(Supplier.name).all(),banks=db.query(Bank).order_by(Bank.name).all(),treasuries=db.query(Treasury).order_by(Treasury.name).all(),filters={"supplier_id":supplier_id,"bank_id":bank_id,"treasury_id":treasury_id,"date_from":date_from or "","date_to":date_to or ""})

@app.get("/supplier-reports/export")
def export_supplier_reports(tab:str="claims",supplier_id:Optional[int]=None,bank_id:Optional[int]=None,treasury_id:Optional[int]=None,date_from:Optional[str]=None,date_to:Optional[str]=None,db:Session=Depends(get_db)):
    tab=tab if tab in SUPPLIER_REPORT_TABS else "claims";columns,rows=supplier_report_data(db,tab,supplier_id,bank_id,treasury_id,date_from,date_to);return journal_export(SUPPLIER_REPORT_TABS[tab],columns,rows,f"AlFarouq_{tab}_Report.xlsx")

@app.get("/health")
def health():
    return {"status": "ok", "app": "صيدليات الفاروق"}
